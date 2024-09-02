import logging
import re

import openpyxl

def read_excel(filename):
    logging.info(f"read_excel filename = {filename}")
    phones = []
    not_correct = []

    try:
        workbook = openpyxl.load_workbook(filename)
        logging.info(f"workbook = {workbook}")
        for sheet in workbook.worksheets:
            logging.info(f"sheet = {sheet}")
            for column in range(1, sheet.max_column + 1):
                logging.debug(f"title cell = {sheet.cell(1, column).value}")
                if str(sheet.cell(1, column).value).lower() == "телефон":
                    for row in range(2, sheet.max_row + 1):
                        logging.debug(f"value cell = {sheet.cell(row, column).value}")
                        phone = sheet.cell(row, column).value
                        clear_phone = re.sub('[() -]', '', str(phone))
                        is_correct_phone = re.match("^\\+7\\d{10}$", clear_phone)
                        if not is_correct_phone:
                            logging.warning(f"phone = {clear_phone} is not a correct phone")
                            not_correct.append(clear_phone)

                        if is_correct_phone and clear_phone not in phones:
                            phones.append(clear_phone)
                    break
    except Exception as ex:
        logging.error("Exception",exc_info=ex)

    return  phones, not_correct