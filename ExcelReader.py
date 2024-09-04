import logging
import re

import openpyxl
from PyQt6.sip import array
from PySide6 import QtCore
from PySide6.QtCore import Signal


class ExcelReader(QtCore.QObject):

    log_signal = Signal(str, str)
    end_signal = Signal(array)

    def read_excel(self, filename):
        logging.info(f"read_excel filename = {filename}")
        phones = []
        not_correct = []
        duplicate = []

        try:
            workbook = openpyxl.load_workbook(filename)
            self.log_signal.emit("excel", "Начата обработка файла " + filename +
                                 ". Количество страниц: " + str(len(workbook.sheetnames)))

            logging.info(f"workbook = {workbook}")
            for sheet in workbook.worksheets:
                logging.info(f"sheet = {sheet}")
                self.log_signal.emit("excel", "Начата обработка страницы " + str(sheet) +
                                     ". Количество строк: " + str(sheet.max_row))

                for column in range(1, sheet.max_column + 1):
                    logging.debug(f"title cell = {sheet.cell(1, column).value}")
                    if str(sheet.cell(1, column).value).lower() == "телефон":
                        for row in range(2, sheet.max_row + 1):
                            logging.debug(f"value cell = {sheet.cell(row, column).value}")
                            phone = sheet.cell(row, column).value
                            clear_phone = re.sub('[() -]', '', str(phone))
                            clear_phone = self.correct_phone(clear_phone)
                            is_correct_phone = re.match("^\\+7\\d{10}$", clear_phone)
                            if not is_correct_phone:
                                logging.warning(f"phone = {clear_phone} is not a correct phone")
                                not_correct.append(clear_phone)

                            if is_correct_phone and clear_phone not in phones:
                                phones.append(clear_phone)
                            elif clear_phone in phones:
                                duplicate.append(clear_phone)
                        break
        except Exception as ex:
            logging.error("Exception",exc_info=ex)
            self.log_signal.emit("excel", "Ошибка при обработке файла " + filename + ": " + str(ex))
            self.end_signal.emit([])
            return

        self.log_signal.emit("excel", "Конец обработки файла " + filename +
                             ". Корректных телефонов: " + str(len(phones)) +
                             ". Некорректных телефонов: " + str(len(not_correct)) +
                             ". Дублей: " + str(len(duplicate)))

        logging.debug(f"read phones: {phones}, not_correct: {not_correct}")

        self.log_signal.emit("excel", "Некорректные телефоны: " + str(not_correct))
        self.end_signal.emit(phones)

    def correct_phone(self, phone):
        need_plus_phone = re.match("^7\\d{10}$", phone)
        if need_plus_phone:
            return "+" + phone

        eight_phone = re.match("^8\\d{10}$", phone)

        if eight_phone:
            return "+7" + phone[1:]
        return phone
